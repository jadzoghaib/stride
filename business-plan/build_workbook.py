"""Generate Stride_Financial_Model.xlsx — a live, formula-driven model.

    uv run python business-plan/build_workbook.py

DESIGN PRINCIPLE: **the workbook contains the logic, not the answers.** Every
cell outside the Assumptions sheet is an Excel formula referring to other cells,
so changing one input on Assumptions recalculates the whole model — including
the three statements and the valuation. Pasting computed values would produce a
report; this produces a model you can interrogate.

Colour convention, stated on the README sheet and used consistently:

    BLUE   an input. Change these.
    BLACK  a formula. Do not overtype — you would break the chain.
    GREEN  a reference to another sheet.

Sheets, in dependency order:

    README        how to use it, and what every colour means
    Assumptions   every input, grouped, with units and provenance
    Drivers       athletes and fans, including cohort churn mechanics
    Revenue       GMV and net revenue, built stream by stream
    Costs         COGS and opex, built line by line
    P&L           income statement
    WorkingCap    receivables, payables, athlete float, capex and D&A
    CashFlow      indirect method, tying to the closing cash balance
    BalanceSheet  with an explicit balance check
    Valuation     DCF, NPV, IRR, exit multiples, WACC/growth sensitivity
    Funding       rounds, dilution, ownership
    Check         the Python model's numbers, to verify the formulas agree
"""

from __future__ import annotations

import pathlib
import re
import sys

sys.path.insert(0, str(pathlib.Path(__file__).parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import market_model as MKT
import model as M
import research_data as RD
import comparables_data as CD

YEARS = M.YEARS
N = len(YEARS)
FIRST = 3                      # column C = Y1
COLS = [get_column_letter(FIRST + k) for k in range(N)]

# ── styling ──────────────────────────────────────────────────────────────────
# Cell FILL carries the meaning, in light tints so a dense sheet stays readable.
# Font colour follows the banking convention as a second signal.
#
#   LIGHT BLUE    an input you may change
#   LIGHT AMBER   sourced external fact (Stripe pricing, Spanish tax law, AWS
#                 list price, Eurobarometer) — change only if the source changed
#   WHITE         a formula computed on this sheet
#   LIGHT GREEN   a formula pulling from another sheet
#   LIGHT GREY    a subtotal or total
#   LIGHT RED     a check that must read zero
FILL_INPUT = PatternFill("solid", fgColor="DCE9F7")
FILL_HARD  = PatternFill("solid", fgColor="FDF0D5")
FILL_CALC  = PatternFill("solid", fgColor="FFFFFF")
FILL_LINK  = PatternFill("solid", fgColor="E4F2E4")
FILL_TOTAL = PatternFill("solid", fgColor="EDF0F5")
FILL_CHECK = PatternFill("solid", fgColor="FBE3E3")

FONT_INPUT = Font(color="1F4E9C", name="Calibri", size=10)
FONT_HARD  = Font(color="8A5200", name="Calibri", size=10)
FONT_CALC  = Font(color="1A1A1A", name="Calibri", size=10)
FONT_LINK  = Font(color="1E7A3C", name="Calibri", size=10)

INPUT = FONT_INPUT
FORMULA = FONT_CALC
LINK = FONT_LINK
BOLD = Font(bold=True, name="Calibri", size=10)
TITLE = Font(bold=True, size=13, color="14100A", name="Calibri")
HEAD = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
HEAD_FILL = PatternFill("solid", fgColor="14181F")
BAND = FILL_TOTAL
ACCENT = PatternFill("solid", fgColor="FFB020")
MONEY = '#,##0;[Red](#,##0)'
MONEY2 = '#,##0.00;[Red](#,##0.00)'
PCT = '0.0%'
NUM = '#,##0'
THIN = Side(style="thin", color="D0D5DD")
TOPLINE = Border(top=Side(style="medium", color="14181F"))
EDGE = Border(left=Side(style="thin", color="D8DEE7"), right=Side(style="thin", color="D8DEE7"),
              top=Side(style="thin", color="D8DEE7"), bottom=Side(style="thin", color="D8DEE7"))

LEGEND = [("Input — change me", FILL_INPUT, FONT_INPUT),
          ("Sourced fact", FILL_HARD, FONT_HARD),
          ("Formula (this sheet)", FILL_CALC, FONT_CALC),
          ("Formula (other sheet)", FILL_LINK, FONT_LINK),
          ("Total", FILL_TOTAL, BOLD)]


def legend(ws, row_ix=2):
    """A compact key on every sheet, so the colours never need explaining."""
    col = 3
    for text, fill, font in LEGEND:
        c = ws.cell(row_ix, col, text)
        rgb = font.color.rgb if font.color is not None else "1A1A1A"
        c.fill, c.font, c.border = fill, Font(size=8, color=rgb, name="Calibri", bold=True), EDGE
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row_ix, start_column=col, end_row=row_ix, end_column=col + 1)
        col += 2


def sheet(wb, name, title):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 15
    for c in COLS:
        ws.column_dimensions[c].width = 13
    ws.freeze_panes = "C4"
    legend(ws)
    r = 3
    ws.cell(r, 1, "Line").font = HEAD
    ws.cell(r, 1).fill = HEAD_FILL
    ws.cell(r, 2, "Unit").font = HEAD
    ws.cell(r, 2).fill = HEAD_FILL
    for k, c in enumerate(COLS):
        cell = ws.cell(r, FIRST + k, f"Y{YEARS[k]}")
        cell.font, cell.fill, cell.alignment = HEAD, HEAD_FILL, Alignment(horizontal="center")
    return ws


def row(ws, r, label, unit="", *, values=None, formula=None, first=None, fmt=MONEY,
        bold=False, font=None, band=False, top=False, indent=0,
        hard=False, const=False, check=False):
    """Write one line.

    `values`   writes inputs (light blue, or amber when `hard`)
    `formula`  writes a template: {c} = this column, {p} = previous, {k} = year
    `const`    a single input in Y1 that later years link to, so a global
               assumption is entered once and visibly propagates
    `hard`     an externally sourced fact rather than one of our estimates

    Fill is chosen from the content: a formula touching another sheet is green,
    a local formula white, a total grey. That way the colour coding cannot drift
    from what the cell actually does.
    """
    lab = ws.cell(r, 1, ("    " * indent) + label)
    lab.font = BOLD if bold else Font(name="Calibri", size=10)
    ws.cell(r, 2, unit).font = Font(name="Calibri", size=9, color="6B7480")

    for k, c in enumerate(COLS):
        cell = ws.cell(r, FIRST + k)
        fill = FILL_CALC
        base = font or FORMULA

        if values is not None:
            if const and k > 0:
                cell.value = f"={COLS[0]}{r}"          # links back to the one input
                fill, base = FILL_LINK, FONT_LINK
            else:
                cell.value = values[k] if k < len(values) else None
                fill = FILL_HARD if hard else FILL_INPUT
                base = FONT_HARD if hard else FONT_INPUT
        elif formula is not None or first is not None:
            prev = COLS[k - 1] if k > 0 else None
            # `first` exists because an accumulating row written as
            # `=IF(k=1, seed, prev + x)` still NAMES its own cell in the branch
            # that never runs, and Excel's dependency graph does not care which
            # branch is taken — it reports a circular reference and stops
            # calculating. Six rows shipped that way. Giving year one its own
            # formula removes the back-reference instead of hiding it.
            src = first if (k == 0 and first is not None) else formula
            f = _quote_sheets(src.format(c=c, p=prev or c, k=k + 1, y=YEARS[k]))
            cell.value = f
            fill = FILL_LINK if "!" in f else FILL_CALC
            base = FONT_LINK if "!" in f else (font or FONT_CALC)

        if band or bold:
            fill = FILL_TOTAL
        if check:
            fill = FILL_CHECK
        cell.fill = fill
        cell.border = EDGE
        cell.number_format = fmt
        cell.font = Font(bold=bold, name="Calibri", size=10,
                         color=(base.color.rgb if base.color else "1A1A1A"))
        if top:
            cell.border = TOPLINE

    if band or bold:
        lab.fill = FILL_TOTAL
        ws.cell(r, 2).fill = FILL_TOTAL
    if top:
        lab.border = TOPLINE
    return r + 1


# Sheet names that are not bare identifiers must be quoted inside a formula.
# `P&L` is the one here: Excel stops reading the sheet token at the ampersand,
# so `=P&L!C21` is a syntax error rather than a reference, and CashFlow,
# BalanceSheet and Check all opened as #NAME?. openpyxl stores the bad string
# happily and a reference-resolution check passes it, which is exactly how forty
# of them shipped — the file only fails when Excel itself parses it. Applied at
# the single point a formula is written, so no future reference can miss it.
_NEEDS_QUOTES = ("P&L",)

# Funding rows, declared once because CashFlow references the equity row before
# the Funding sheet is built. The builder asserts the rows land here, so the
# declaration cannot quietly stop being true — the previous version hand-counted
# them, was off by one throughout, and made every Funding formula circular while
# CashFlow booked a pre-money valuation as cash received.
FUNDING_ROWS = {
    "Equity raised": 4,
    "Pre-money valuation": 5,
    "Post-money valuation": 6,
    "New investor stake": 7,
    "Cumulative dilution": 8,
    "Founders + team retained": 9,
}


def _quote_sheets(formula: str) -> str:
    for name in _NEEDS_QUOTES:
        formula = re.sub(rf"(?<!')\b{re.escape(name)}!", f"'{name}'!", formula)
    return formula


def section(ws, r, text):
    c = ws.cell(r, 1, text)
    c.font = Font(bold=True, size=10, color="8A5200")
    c.fill = PatternFill("solid", fgColor="FFF4DE")
    for k in range(len(COLS) + 1):
        ws.cell(r, 2 + k).fill = PatternFill("solid", fgColor="FFF4DE")
    return r + 1


def build() -> pathlib.Path:
    wb = Workbook()
    wb.remove(wb.active)
    A = M.A
    rows = M.build()

    # ══ README ══════════════════════════════════════════════════════════════
    rd = wb.create_sheet("README")
    rd.column_dimensions["A"].width = 100
    lines = [
        ("Stride — Financial Model", TITLE),
        ("", None),
        ("10-year operating model, three statements, and a valuation. EUR. Y1 = 2027.", BOLD),
        ("", None),
        ("HOW TO USE", BOLD),
        ("Change any BLUE cell on the Assumptions sheet. Everything else recalculates.", None),
        ("Every non-input cell is a formula, so you can trace any number back to the inputs", None),
        ("that produced it — select a cell and use Formulas > Trace Precedents.", None),
        ("", None),
        ("COLOURS — the fill tells you what a cell is", BOLD),
        ("  LIGHT BLUE    an input. Change these. Blue is our estimate or plan.", None),
        ("  LIGHT AMBER   a sourced fact — Stripe pricing, Spanish tax law, AWS list,", None),
        ("                Eurobarometer. Change only if the source changed.", None),
        ("  WHITE         a formula computed on this sheet.", None),
        ("  LIGHT GREEN   a formula pulling a value from another sheet.", None),
        ("  LIGHT GREY    a subtotal or total.", None),
        ("  LIGHT RED     a check that must read zero.", None),
        ("", None),
        ("A global assumption is entered once, in the Y1 column, and later years link", None),
        ("back to it in green — so you change one cell and the whole row follows. Where", None),
        ("a row genuinely varies year by year, every cell is blue.", None),
        ("", None),
        ("The workbook is set to recalculate on open, so every figure you see was", None),
        ("computed from the inputs rather than pasted.", None),
        ("", None),
        ("SHEETS", BOLD),
        ("  Assumptions    every input, grouped, with units and provenance", None),
        ("  Comparables    published facts about OnlyFans, Patreon, Passes and agents", None),
        ("  MarketModel    turns those facts into our assumptions, step by step", None),
        ("  Research       how each assumption was baselined, with sources", None),
        ("  Drivers        athletes and fans, including monthly cohort churn", None),
        ("  Revenue        GMV and net revenue, stream by stream", None),
        ("  Costs          COGS and operating costs, line by line", None),
        ("  P&L            income statement", None),
        ("  WorkingCap     receivables, payables, athlete float, capex, D&A", None),
        ("  CashFlow       indirect method, tying to closing cash", None),
        ("  BalanceSheet   with an explicit balance check row", None),
        ("  Valuation      DCF, NPV, IRR, exit multiples, sensitivity", None),
        ("  Funding        rounds, dilution, ownership", None),
        ("  Check          the Python model's figures, to verify the formulas agree", None),
        ("", None),
        ("WHY TEN YEARS", BOLD),
        ("At Y7 the business is still compounding above 50%, so a terminal value placed", None),
        ("there does most of the valuation work and does it badly. Ten years lets growth", None),
        ("decelerate inside the explicit forecast, where it can be argued with.", None),
        ("", None),
        ("EVIDENCE CHAIN", BOLD),
        ("  Comparables (published facts) -> MarketModel (derives ours) -> Assumptions", None),
        ("  -> Drivers -> Revenue and Costs -> statements -> Valuation.", None),
        ("  From Assumptions rightwards, every link is a formula you can follow with", None),
        ("  Trace Precedents. The first arrow is checked rather than linked: MarketModel", None),
        ("  derives each figure from the published ones, and Assumptions carries the", None),
        ("  result rounded for reading -- 37 fans per athlete, not 37.026. Linking the", None),
        ("  cells would put the workbook a hair off the Python everywhere, which the", None),
        ("  Check sheet exists to forbid, so scripts/doc_consistency.py asserts the", None),
        ("  derivation still rounds to those literals instead. Refresh a comparable far", None),
        ("  enough to change one of them at the precision it is written and that check", None),
        ("  fails. A refresh too small to change any displayed figure passes, which is", None),
        ("  the same statement, not a loophole in it.", None),
        ("", None),
        ("MODEL SHAPE", BOLD),
        ("This is a target-driven model: athlete counts are the plan, and marketing spend", None),
        ("is derived from them at segment CAC. It is not a driver-driven model in which", None),
        ("spend produces athletes. That is the normal shape for a plan, but it means the", None),
        ("athlete trajectory is an assumption to defend, not an output to trust.", None),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        c = rd.cell(i, 1, text)
        if font:
            c.font = font

    # ══ ASSUMPTIONS ═════════════════════════════════════════════════════════
    a = sheet(wb, "Assumptions", "Assumptions — every input lives here")
    r = 4
    A_ROW = {}

    def put(label, unit, values, fmt=NUM, key=None, hard=False, const=False, note="",
            derive_note=""):
        nonlocal r
        # An input that repeats a number already on `model.A` instead of reading
        # it is a divergence waiting to happen: retune the Python and the
        # workbook goes on costing the old plan, silently, because both files
        # still parse. Wherever the key names a real field, the two must agree.
        if key and hasattr(A, key):
            expected = getattr(A, key)
            if isinstance(expected, (int, float)) and not isinstance(expected, bool):
                assert all(abs(v - expected) < 1e-9 for v in values), (
                    f"Assumptions row {label!r} writes {values[0]} but model.A.{key} "
                    f"is {expected} — the workbook and the Python have diverged")
            elif isinstance(expected, list):
                assert all(abs(a - b) < 1e-9 for a, b in zip(values, expected)), (
                    f"Assumptions row {label!r} does not match model.A.{key}")
        A_ROW[key or label] = r
        r = row(a, r, label, unit, values=values, fmt=fmt, hard=hard, const=const)
        if note or derive_note:
            c = a.cell(r - 1, FIRST + N + 1, note or derive_note)
            c.font = Font(size=8, italic=True,
                          color="1E7A3C" if derive_note else "6B7480", name="Calibri")

    r = section(a, r, "MARKET — the plan's shape")
    put("Active athletes (year end)", "count", A.athletes, NUM, "athletes")
    put("Niche share of athletes", "%", A.niche_share, PCT, "niche_share")
    put("Athletes needed for full SaaS value", "count", [A.athletes_for_full_saas_value] * N,
        NUM, "saas_floor", const=True,
        note="Below this, sponsor conversion scales with supply — a matching product "
             "is worth nothing against an empty directory")
    r += 1

    for seg, tag in ((M.NICHE, "niche"), (M.POPULAR, "popular")):
        r = section(a, r, f"SEGMENT — {tag.upper()}")
        put(f"Monetising athletes ({tag})", "% of segment", seg.monetise_rate, PCT, f"{tag}_monetise")
        put(f"Paying fans per monetising athlete ({tag})", "count", seg.fans_per_athlete, NUM, f"{tag}_fpa",
            derive_note="derived on MarketModel from Patreon members-per-creator")
        put(f"Fan ARPU per month ({tag})", "EUR", seg.fan_arpu_month, MONEY2, f"{tag}_arpu",
            derive_note="derived on MarketModel from our tier mix, cross-checked against Patreon")
        put(f"Fan churn per month ({tag})", "%", seg.fan_churn_month, PCT, f"{tag}_fchurn",
            derive_note="derived on MarketModel from Patreon churn and our annual-plan mix")
        put(f"Athlete churn per year ({tag})", "%", seg.athlete_churn_year, PCT, f"{tag}_achurn")
        put(f"Max new fans per athlete per year ({tag})", "capacity ceiling",
            seg.max_fan_adds_per_athlete_year, NUM, f"{tag}_maxadds",
            note="An audience is finite; without this ceiling higher churn RAISES revenue")
        put(f"Athletes landing a deal ({tag})", "% of segment", seg.deal_rate, PCT, f"{tag}_dealrate")
        put(f"Deals per dealing athlete ({tag})", "count", seg.deals_per_athlete, '0.0', f"{tag}_dpa")
        put(f"Average deal value ({tag})", "EUR", seg.avg_deal_eur, MONEY, f"{tag}_deal")
        put(f"Athlete CAC ({tag})", "EUR", seg.cac_eur, MONEY, f"{tag}_cac")
        r += 1

    r = section(a, r, "SPONSORS")
    put("Registered sponsors", "count", A.sponsors, NUM, "sponsors")
    put("Share on a paid plan", "%", A.sponsor_paid_rate, PCT, "sponsor_paid")
    put("Sponsor plan ARPU per month", "EUR", A.sponsor_arpu_month, MONEY, "sponsor_arpu")
    put("Sponsor CAC", "EUR", A.sponsor_cac_eur, MONEY, "sponsor_cac")
    r += 1

    r = section(a, r, "TAKE RATES & PAYMENT RAILS")
    put("Take rate — fan revenue", "%", [A.take_fan] * N, PCT, "take_fan", const=True,
        note="Our pricing decision. OnlyFans/Fansly/Fanfix 20%, Passes 10% + $29/mo")
    put("Take rate — sponsorship", "%", [A.take_sponsorship] * N, PCT, "take_sp", const=True,
        note="Agents take 10-20% of an endorsement")
    put("PPV & tips as multiple of subs", "x", [A.ppv_tips_multiple] * N, '0.00', "ppv_mult", const=True)
    put("PSP percentage fee", "%", [A.psp_pct] * N, '0.00%', "psp_pct", hard=True, const=True,
        note="Stripe published pricing")
    put("PSP fixed fee per transaction", "EUR", [A.psp_fixed_eur] * N, MONEY2, "psp_fix", hard=True, const=True,
        note="Stripe. The single most damaging cost at small ticket sizes")
    put("Payout percentage fee", "%", [A.payout_pct] * N, '0.00%', "payout_pct", hard=True, const=True,
        note="Stripe Connect")
    put("Payout fixed fee", "EUR", [A.payout_fixed_eur] * N, MONEY2, "payout_fix", hard=True, const=True,
        note="Stripe Connect")
    put("Average fan transaction", "EUR", [A.avg_fan_txn_eur] * N, MONEY2, "fan_txn", const=True)
    put("Average deal transaction", "EUR", [A.avg_deal_txn_eur] * N, MONEY, "deal_txn", const=True)
    r += 1

    r = section(a, r, "INFRASTRUCTURE & CONTENT COSTS")
    put("AWS base cost per month", "EUR", A.aws_base_month, MONEY, "aws")
    put("Media GB per paying fan per month", "GB", [A.gb_per_fan_month] * N, '0.0', "gb", const=True)
    put("Egress cost per GB (zero-egress CDN)", "EUR", [A.egress_eur_per_gb] * N, '0.000', "egress", hard=True, const=True,
        note="Cloudflare R2 / Backblaze B2 list")
    put("Egress cost per GB (CloudFront list)", "EUR", [A.egress_eur_per_gb_naive] * N, '0.000', "egress_naive", hard=True, const=True,
        note="AWS CloudFront list price — the EUR 1.1M/yr trap")
    put("Moderation cost per 1,000 items", "EUR", [A.moderation_eur_per_1k_items] * N, MONEY2, "mod_rate", const=True)
    put("Items per athlete per month", "count", [A.items_per_athlete_month] * N, '0.0', "items", const=True)
    r += 1

    r = section(a, r, "PEOPLE & OVERHEAD")
    put("Headcount", "FTE", A.headcount, '0.0', "headcount")
    put("Loaded salary (incl. ~31% employer SS)", "EUR", A.loaded_salary_eur, MONEY, "salary")
    put("Legal & compliance", "EUR", A.legal_compliance_eur, MONEY, "legal")
    put("Other opex as % of revenue", "%", [A.other_opex_pct_of_revenue] * N, PCT, "other_pct", const=True)
    r += 1

    r = section(a, r, "WORKING CAPITAL, CAPEX & TAX")
    put("Athlete payout float", "days", [A.float_days] * N, NUM, "float_days", const=True,
        note="We hold fan money before paying athletes — a cash benefit")
    put("Sponsor receivable days", "days", [A.ar_days] * N, NUM, "ar_days", const=True)
    put("Trade payable days", "days", [A.ap_days] * N, NUM, "ap_days", const=True)
    put("Capitalised development", "% of people cost", [A.capex_pct] * N, PCT, "capex_pct",
        const=True)
    put("Amortisation period", "years", [A.amort_years] * N, '0', "amort_years", const=True)
    put("Corporate tax — startup rate", "%", [A.tax_low] * N, PCT, "tax_low", hard=True, const=True,
        note="Spanish Startup Law, first 4 profitable years")
    put("Corporate tax — standard rate", "%", [A.tax_high] * N, PCT, "tax_high", hard=True, const=True,
        note="Spanish corporate income tax")
    put("Years at startup rate (from first profit)", "years", [A.tax_low_years] * N, '0',
        "tax_low_years", hard=True, const=True)
    r += 1

    r = section(a, r, "ADMISSION — the gate between an applicant and an athlete")
    put("Admission rate — direct applicants", "% admitted", [A.admission_rate_direct] * N, PCT,
        "admit_direct", const=True,
        derive_note="ops-load output of scripts/admission_stress.py")
    put("Review rate — direct applicants", "% needing a human", [A.review_rate_direct] * N, PCT,
        "review_direct", const=True,
        derive_note="ops-load output of scripts/admission_stress.py")
    put("Admission rate — club-nominated", "% admitted", [A.admission_rate_club] * N, PCT,
        "admit_club", const=True,
        note="ESTIMATE — a verified club's floor carries more of them past the bar")
    put("Review rate — club-nominated", "% needing a human", [A.review_rate_club] * N, PCT,
        "review_club", const=True, note="ESTIMATE")
    put("Applicants arriving via a club", "% of applicants", A.club_sourced_share, PCT, "club_share",
        note="Grows as federation and club partnerships land")
    put("Minutes per manual review", "min", [A.review_minutes] * N, '0.0', "review_min", const=True,
        note="Open the link, decide whether it names the applicant")
    put("Productive hours per reviewer-year", "h", [A.ops_hours_per_year] * N, NUM, "ops_hours",
        const=True, note="Prices review time off the same loaded salary line")
    r += 1

    r = section(a, r, "VALUATION")
    put("WACC / discount rate", "%", [A.wacc] * N, PCT, "wacc", const=True,
        note="Early-stage venture hurdle")
    put("Terminal growth", "%", [A.terminal_growth] * N, PCT, "tg", const=True)
    put("Exit revenue multiple", "x", [6.5] * N, '0.0', "exit_mult", const=True,
        note="Blended marketplace + SaaS comparables")
    put("Risk-free rate (Spanish 10Y)", "%", [A.risk_free] * N, PCT, "rf", hard=True, const=True,
        note="Spanish 10Y government bond, mid-2026")
    put("Founder alternative return", "%", [A.founder_alt_return] * N, PCT, "alt", const=True)

    def AR(key, col):
        return f"Assumptions!${col}${A_ROW[key]}"

    def A1(key):
        """First-column absolute reference (for scalars repeated across years)."""
        return f"Assumptions!$C${A_ROW[key]}"

    # ══ DRIVERS ═════════════════════════════════════════════════════════════
    d = sheet(wb, "Drivers", "Drivers — athletes and fans, with cohort churn")
    r, D = 4, {}

    def drow(label, formula=None, values=None, fmt=NUM, **kw):
        nonlocal r
        D[label] = r
        r = row(d, r, label, kw.pop("unit", ""), formula=formula, values=values, fmt=fmt, **kw)

    r = section(d, r, "ATHLETES")
    drow("Total athletes (year end)", formula=f"={{c}}{A_ROW['athletes']}".replace("{c}", "Assumptions!{c}"), fmt=NUM, font=LINK)
    drow("Niche athletes", formula="=Drivers!{c}" + str(D["Total athletes (year end)"]) + "*Assumptions!{c}" + str(A_ROW["niche_share"]))
    drow("Popular athletes", formula="=Drivers!{c}" + str(D["Total athletes (year end)"]) + "-Drivers!{c}" + str(D["Niche athletes"]))
    r += 1

    for tag, base in (("Niche", "Niche athletes"), ("Popular", "Popular athletes")):
        t = tag.lower()
        r = section(d, r, f"{tag.upper()} — fans and churn")
        drow(f"{tag} athletes lost to churn",
             formula=(f"=IF({{k}}=1,0,{{p}}{D[base]}*Assumptions!{{c}}{A_ROW[f'{t}_achurn']})"))
        drow(f"{tag} athletes acquired (gross)",
             formula=(f"=MAX(0,{{c}}{D[base]}-IF({{k}}=1,0,{{p}}{D[base]})+{{c}}{r-1})"))
        drow(f"{tag} monetising athletes",
             formula=f"={{c}}{D[base]}*Assumptions!{{c}}{A_ROW[f'{t}_monetise']}")
        drow(f"{tag} paying fans (year end)",
             formula=f"={{c}}{r-1}*Assumptions!{{c}}{A_ROW[f'{t}_fpa']}", bold=True)
        # cohort mechanics, spelled out so the maths is visible rather than asserted
        drow(f"{tag} monthly retention r = 1-churn", unit="factor",
             formula=f"=1-Assumptions!{{c}}{A_ROW[f'{t}_fchurn']}", fmt='0.000')
        drow(f"{tag} opening fans survived 12 months", unit="= open x r^12",
             formula=f"=IF({{k}}=1,0,{{p}}{D[f'{tag} paying fans (year end)']}*{{c}}{r-1}^12)")
        rr = D[f"{tag} monthly retention r = 1-churn"]
        surv = r - 1
        # Capped HERE, on the driver, because everything downstream reads this
        # row: the annual figure below it and — the one that matters — average
        # fans during the year, which is what revenue accrues on. Capping only
        # the annual display row left the ceiling cosmetic and the revenue
        # uncapped, which is the exact shape of the bug the ceiling was added to
        # prevent. Python caps `monthly_adds` in fan_path(); this is that.
        drow(f"{tag} monthly gross adds", unit="solved, then capped by capacity",
             formula=(f"=MIN(MAX(0,({{c}}{D[f'{tag} paying fans (year end)']}-{{c}}{surv})"
                      f"/((1-{{c}}{rr}^12)/(1-{{c}}{rr}))),"
                      f"{{c}}{D[f'{tag} monetising athletes']}"
                      f"*Assumptions!{{c}}{A_ROW[f'{t}_maxadds']}/12)"))
        # Plain x12 again: the ceiling is applied to the monthly row above, so
        # this inherits it rather than clamping a second time.
        drow(f"{tag} fans acquired (gross)", unit="x12 months",
             formula=f"={{c}}{r-1}*12")
        drow(f"{tag} S = r(1-r^12)/(1-r)", unit="geometric sum", fmt='0.000',
             formula=f"={{c}}{rr}*(1-{{c}}{rr}^12)/(1-{{c}}{rr})")
        drow(f"{tag} average fans during year", unit="exact mean, revenue basis",
             formula=(f"=(IF({{k}}=1,0,{{p}}{D[f'{tag} paying fans (year end)']})*{{c}}{r-1}"
                      f"+{{c}}{D[f'{tag} monthly gross adds']}/(1-{{c}}{rr})*(12-{{c}}{r-1}))/12"),
             bold=True)
        drow(f"{tag} fans lost to churn", unit="avg x churn x 12",
             formula=(f"={{c}}{D[f'{tag} average fans during year']}*12"
                      f"*Assumptions!{{c}}{A_ROW[f'{t}_fchurn']}"))
        r += 1

    r = section(d, r, "TOTALS")
    drow("Paying fans (year end)",
         formula=f"={{c}}{D['Niche paying fans (year end)']}+{{c}}{D['Popular paying fans (year end)']}", bold=True)
    drow("Average paying fans",
         formula=f"={{c}}{D['Niche average fans during year']}+{{c}}{D['Popular average fans during year']}", bold=True)
    drow("Athletes acquired (gross)",
         formula=f"={{c}}{D['Niche athletes acquired (gross)']}+{{c}}{D['Popular athletes acquired (gross)']}")
    drow("Fans lost to churn",
         formula=f"={{c}}{D['Niche fans lost to churn']}+{{c}}{D['Popular fans lost to churn']}")
    r += 1

    r = section(d, r, "DEALS")
    drow("Niche deals",
         formula=(f"={{c}}{D['Niche athletes']}*Assumptions!{{c}}{A_ROW['niche_dealrate']}"
                  f"*Assumptions!{{c}}{A_ROW['niche_dpa']}"))
    drow("Popular deals",
         formula=(f"={{c}}{D['Popular athletes']}*Assumptions!{{c}}{A_ROW['popular_dealrate']}"
                  f"*Assumptions!{{c}}{A_ROW['popular_dpa']}"))
    drow("Total deals", formula=f"={{c}}{r-2}+{{c}}{r-1}", bold=True)
    # A matching product is worth nothing without athletes to match. Below the
    # supply floor, paid conversion scales with the size of the directory —
    # otherwise the workbook books SaaS revenue against an empty one, and again
    # disagrees with Python.
    drow("Supply factor", unit="athletes / floor, capped at 1", fmt='0.000',
         formula=(f"=MIN(1,{{c}}{D['Total athletes (year end)']}"
                  f"/Assumptions!{{c}}{A_ROW['saas_floor']})"))
    drow("Paying sponsors",
         formula=(f"=Assumptions!{{c}}{A_ROW['sponsors']}"
                  f"*Assumptions!{{c}}{A_ROW['sponsor_paid']}"
                  f"*{{c}}{D['Supply factor']}"))
    drow("Sponsors acquired (gross)",
         formula=f"=MAX(0,Assumptions!{{c}}{A_ROW['sponsors']}-IF({{k}}=1,0,Assumptions!{{p}}{A_ROW['sponsors']}))")
    r += 1

    r = section(d, r, "ADMISSION FUNNEL")
    drow("Blended admission rate", unit="weighted by channel", fmt=PCT,
         formula=(f"=Assumptions!{{c}}{A_ROW['club_share']}*Assumptions!{{c}}{A_ROW['admit_club']}"
                  f"+(1-Assumptions!{{c}}{A_ROW['club_share']})*Assumptions!{{c}}{A_ROW['admit_direct']}"))
    drow("Blended review rate", unit="weighted by channel", fmt=PCT,
         formula=(f"=Assumptions!{{c}}{A_ROW['club_share']}*Assumptions!{{c}}{A_ROW['review_club']}"
                  f"+(1-Assumptions!{{c}}{A_ROW['club_share']})*Assumptions!{{c}}{A_ROW['review_direct']}"))
    drow("Applications required", unit="gross adds / admission rate",
         formula=(f"={{c}}{D['Athletes acquired (gross)']}"
                  f"/{{c}}{D['Blended admission rate']}"))
    drow("Manual reviews", unit="applications x review rate",
         formula=f"={{c}}{D['Applications required']}*{{c}}{D['Blended review rate']}")
    drow("Reviewer FTE implied", unit="the real constraint, not the euros", fmt='0.00',
         formula=(f"={{c}}{D['Manual reviews']}*Assumptions!{{c}}{A_ROW['review_min']}/60"
                  f"/Assumptions!{{c}}{A_ROW['ops_hours']}"))

    # ══ REVENUE ═════════════════════════════════════════════════════════════
    v = sheet(wb, "Revenue", "Revenue — GMV built stream by stream, then our take")
    r, R = 4, {}

    def vrow(label, formula=None, fmt=MONEY, **kw):
        nonlocal r
        R[label] = r
        r = row(v, r, label, kw.pop("unit", ""), formula=formula, fmt=fmt, **kw)

    r = section(v, r, "FAN GMV — what fans pay athletes")
    vrow("Niche subscription GMV", unit="avg fans x ARPU x 12",
         formula=(f"=Drivers!{{c}}{D['Niche average fans during year']}"
                  f"*Assumptions!{{c}}{A_ROW['niche_arpu']}*12"))
    vrow("Popular subscription GMV",
         formula=(f"=Drivers!{{c}}{D['Popular average fans during year']}"
                  f"*Assumptions!{{c}}{A_ROW['popular_arpu']}*12"))
    vrow("Subscription GMV", formula=f"={{c}}{r-2}+{{c}}{r-1}", bold=True)
    vrow("PPV and tips GMV", unit="x multiple",
         formula=f"={{c}}{R['Subscription GMV']}*Assumptions!{{c}}{A_ROW['ppv_mult']}")
    vrow("Total fan GMV", formula=f"={{c}}{R['Subscription GMV']}+{{c}}{R['PPV and tips GMV']}",
         bold=True, band=True)
    r += 1

    r = section(v, r, "SPONSORSHIP GMV — what sponsors pay athletes and clubs")
    vrow("Niche sponsorship GMV",
         formula=f"=Drivers!{{c}}{D['Niche deals']}*Assumptions!{{c}}{A_ROW['niche_deal']}")
    vrow("Popular sponsorship GMV",
         formula=f"=Drivers!{{c}}{D['Popular deals']}*Assumptions!{{c}}{A_ROW['popular_deal']}")
    vrow("Total sponsorship GMV", formula=f"={{c}}{r-2}+{{c}}{r-1}", bold=True, band=True)
    vrow("TOTAL GMV", formula=f"={{c}}{R['Total fan GMV']}+{{c}}{R['Total sponsorship GMV']}",
         bold=True, top=True)
    r += 1

    r = section(v, r, "NET REVENUE — our share")
    vrow("Fan take", unit="fan GMV x take rate",
         formula=f"={{c}}{R['Total fan GMV']}*Assumptions!{{c}}{A_ROW['take_fan']}")
    vrow("Sponsorship take",
         formula=f"={{c}}{R['Total sponsorship GMV']}*Assumptions!{{c}}{A_ROW['take_sp']}")
    vrow("Sponsor SaaS", unit="paying sponsors x ARPU x 12",
         formula=(f"=Drivers!{{c}}{D['Paying sponsors']}"
                  f"*Assumptions!{{c}}{A_ROW['sponsor_arpu']}*12"))
    vrow("NET REVENUE", formula=f"=SUM({{c}}{r-3}:{{c}}{r-1})", bold=True, top=True, band=True)
    vrow("Growth", unit="% YoY", formula=f"=IF({{k}}=1,\"\",{{c}}{r-1}/{{p}}{r-1}-1)", fmt=PCT)
    vrow("Take rate on GMV", unit="blended",
         formula=f"={{c}}{R['NET REVENUE']}/{{c}}{R['TOTAL GMV']}", fmt=PCT)

    # ══ COSTS ═══════════════════════════════════════════════════════════════
    co = sheet(wb, "Costs", "Costs — every line built from its driver")
    r, C = 4, {}

    def crow(label, formula=None, fmt=MONEY, **kw):
        nonlocal r
        C[label] = r
        r = row(co, r, label, kw.pop("unit", ""), formula=formula, fmt=fmt, **kw)

    r = section(co, r, "COST OF SALES")
    crow("Fan transactions", unit="fan GMV / avg ticket", fmt=NUM,
         formula=f"=Revenue!{{c}}{R['Total fan GMV']}/Assumptions!{{c}}{A_ROW['fan_txn']}")
    crow("Deal transactions", unit="count", fmt=NUM,
         formula=f"=Revenue!{{c}}{R['Total sponsorship GMV']}/Assumptions!{{c}}{A_ROW['deal_txn']}")
    crow("Payment processing", unit="% of GMV + fixed",
         formula=(f"=Revenue!{{c}}{R['TOTAL GMV']}*Assumptions!{{c}}{A_ROW['psp_pct']}"
                  f"+({{c}}{C['Fan transactions']}+{{c}}{C['Deal transactions']})"
                  f"*Assumptions!{{c}}{A_ROW['psp_fix']}"))
    crow("Payouts to athletes", unit="% + fixed",
         formula=(f"=Revenue!{{c}}{R['TOTAL GMV']}*Assumptions!{{c}}{A_ROW['payout_pct']}"
                  f"+({{c}}{C['Fan transactions']}/30+{{c}}{C['Deal transactions']})"
                  f"*Assumptions!{{c}}{A_ROW['payout_fix']}"))
    crow("Media egress GB", unit="GB/yr", fmt=NUM,
         formula=(f"=Drivers!{{c}}{D['Average paying fans']}"
                  f"*Assumptions!{{c}}{A_ROW['gb']}*12"))
    crow("Infrastructure (AWS + CDN)", unit="base + egress",
         formula=(f"=Assumptions!{{c}}{A_ROW['aws']}*12"
                  f"+{{c}}{C['Media egress GB']}*Assumptions!{{c}}{A_ROW['egress']}"))
    crow("  memo: same egress at CloudFront list", unit="the trap",
         formula=(f"=Assumptions!{{c}}{A_ROW['aws']}*12"
                  f"+{{c}}{C['Media egress GB']}*Assumptions!{{c}}{A_ROW['egress_naive']}"),
         font=Font(color="B3272D", name="Calibri", size=10, italic=True))
    crow("Moderation", unit="items x rate",
         formula=(f"=Drivers!{{c}}{D['Total athletes (year end)']}"
                  f"*Assumptions!{{c}}{A_ROW['items']}*12/1000"
                  f"*Assumptions!{{c}}{A_ROW['mod_rate']}"))
    crow("Athlete verification", unit="reviews x minutes x hourly",
         formula=(f"=Drivers!{{c}}{D['Manual reviews']}"
                  f"*Assumptions!{{c}}{A_ROW['review_min']}/60"
                  f"*(Assumptions!{{c}}{A_ROW['salary']}/Assumptions!{{c}}{A_ROW['ops_hours']})"))
    crow("TOTAL COST OF SALES", bold=True, top=True, band=True,
         formula=(f"={{c}}{C['Payment processing']}+{{c}}{C['Payouts to athletes']}"
                  f"+{{c}}{C['Infrastructure (AWS + CDN)']}+{{c}}{C['Moderation']}"
                  f"+{{c}}{C['Athlete verification']}"))
    r += 1

    r = section(co, r, "OPERATING COSTS")
    crow("People", unit="FTE x loaded salary",
         formula=f"=Assumptions!{{c}}{A_ROW['headcount']}*Assumptions!{{c}}{A_ROW['salary']}")
    crow("Athlete acquisition — niche", unit="gross adds x CAC",
         formula=(f"=Drivers!{{c}}{D['Niche athletes acquired (gross)']}"
                  f"*Assumptions!{{c}}{A_ROW['niche_cac']}"))
    crow("Athlete acquisition — popular",
         formula=(f"=Drivers!{{c}}{D['Popular athletes acquired (gross)']}"
                  f"*Assumptions!{{c}}{A_ROW['popular_cac']}"))
    crow("Sponsor acquisition",
         formula=(f"=Drivers!{{c}}{D['Sponsors acquired (gross)']}"
                  f"*Assumptions!{{c}}{A_ROW['sponsor_cac']}"))
    crow("Total marketing / CAC", bold=True,
         formula=f"=SUM({{c}}{r-3}:{{c}}{r-1})")
    # The ambiguity worth surfacing rather than resolving silently: segment CAC
    # is priced per athlete ADMITTED, so multiplying it by the funnel would
    # double-count. This memo is what the same money works out to per
    # application — the figure to hold against what a channel actually charges.
    crow("  memo: athlete marketing per application", unit="CAC is per ADMITTED athlete",
         fmt=MONEY2, font=Font(color="6B7480", name="Calibri", size=10, italic=True),
         formula=(f"=({{c}}{C['Athlete acquisition — niche']}+{{c}}{C['Athlete acquisition — popular']})"
                  f"/Drivers!{{c}}{D['Applications required']}"))
    crow("Legal & compliance", formula=f"=Assumptions!{{c}}{A_ROW['legal']}", font=LINK)
    crow("Other opex", unit="% of revenue",
         formula=f"=Revenue!{{c}}{R['NET REVENUE']}*Assumptions!{{c}}{A_ROW['other_pct']}")
    crow("TOTAL OPERATING COSTS", bold=True, top=True, band=True,
         formula=(f"={{c}}{C['People']}+{{c}}{C['Total marketing / CAC']}"
                  f"+{{c}}{C['Legal & compliance']}+{{c}}{C['Other opex']}"))

    # ══ WORKING CAPITAL & CAPEX ═════════════════════════════════════════════
    wc = sheet(wb, "WorkingCap", "Working capital, capex and amortisation")
    r, W = 4, {}

    def wrow(label, formula=None, fmt=MONEY, **kw):
        nonlocal r
        W[label] = r
        r = row(wc, r, label, kw.pop("unit", ""), formula=formula, fmt=fmt, **kw)

    r = section(wc, r, "WORKING CAPITAL")
    wrow("Sponsor receivables", unit="SaaS+take x AR days/365",
         formula=(f"=(Revenue!{{c}}{R['Sponsorship take']}+Revenue!{{c}}{R['Sponsor SaaS']})"
                  f"*Assumptions!{{c}}{A_ROW['ar_days']}/365"))
    wrow("Athlete payout float", unit="GMV owed, not yet paid",
         formula=(f"=Revenue!{{c}}{R['TOTAL GMV']}*Assumptions!{{c}}{A_ROW['float_days']}/365"))
    wrow("Trade payables", unit="opex x AP days/365",
         formula=(f"=Costs!{{c}}{C['TOTAL OPERATING COSTS']}*Assumptions!{{c}}{A_ROW['ap_days']}/365"))
    wrow("Net working capital", unit="AR - float - AP", bold=True,
         formula=f"={{c}}{W['Sponsor receivables']}-{{c}}{W['Athlete payout float']}-{{c}}{W['Trade payables']}")
    wrow("Change in NWC", unit="cash impact",
         formula=f"=IF({{k}}=1,{{c}}{W['Net working capital']},{{c}}{W['Net working capital']}-{{p}}{W['Net working capital']})")
    r += 1

    r = section(wc, r, "CAPEX & AMORTISATION")
    wrow("Capitalised development", unit="% of people cost",
         formula=f"=Costs!{{c}}{C['People']}*Assumptions!{{c}}{A_ROW['capex_pct']}")
    wrow("Amortisation", unit="3-yr straight line on trailing capex",
         formula=(f"=SUM(OFFSET({{c}}{W['Capitalised development']},0,"
                  f"-MIN(Assumptions!{{c}}{A_ROW['amort_years']}-1,{{k}}-1),1,"
                  f"MIN(Assumptions!{{c}}{A_ROW['amort_years']},{{k}})))"
                  f"/Assumptions!{{c}}{A_ROW['amort_years']}"))
    wrow("Net intangible assets", unit="closing",
         first=f"={{c}}{W['Capitalised development']}-{{c}}{W['Amortisation']}",
         formula=(f"={{p}}{r}+{{c}}{W['Capitalised development']}"
                  f"-{{c}}{W['Amortisation']}"))

    # ══ P&L ═════════════════════════════════════════════════════════════════
    pl = sheet(wb, "P&L", "Income statement")
    r, P = 4, {}

    def prow(label, formula=None, fmt=MONEY, **kw):
        nonlocal r
        P[label] = r
        r = row(pl, r, label, kw.pop("unit", ""), formula=formula, fmt=fmt, **kw)

    prow("Net revenue", formula=f"=Revenue!{{c}}{R['NET REVENUE']}", font=LINK, bold=True)
    prow("Cost of sales", formula=f"=-Costs!{{c}}{C['TOTAL COST OF SALES']}", font=LINK)
    prow("GROSS PROFIT", formula=f"={{c}}{P['Net revenue']}+{{c}}{P['Cost of sales']}", bold=True, top=True)
    prow("Gross margin", formula=f"={{c}}{P['GROSS PROFIT']}/{{c}}{P['Net revenue']}", fmt=PCT)
    r += 1
    prow("Operating costs", formula=f"=-Costs!{{c}}{C['TOTAL OPERATING COSTS']}", font=LINK)
    prow("EBITDA", formula=f"={{c}}{P['GROSS PROFIT']}+{{c}}{P['Operating costs']}", bold=True, top=True, band=True)
    prow("EBITDA margin", formula=f"={{c}}{P['EBITDA']}/{{c}}{P['Net revenue']}", fmt=PCT)
    prow("Amortisation", formula=f"=-WorkingCap!{{c}}{W['Amortisation']}", font=LINK)
    prow("EBIT", formula=f"={{c}}{P['EBITDA']}+{{c}}{P['Amortisation']}", bold=True)
    r += 1
    r = section(pl, r, "TAX — with loss carryforward and the startup-rate step")
    prow("Losses brought forward", unit="accumulated",
         formula=f"=IF({{k}}=1,0,MIN(0,{{p}}{r+2}))")
    prow("Taxable profit", unit="after offset",
         formula=f"=MAX(0,{{c}}{P['EBIT']}+{{c}}{r-1})")
    prow("Losses carried forward", unit="running",
         first=f"=MIN(0,{{c}}{P['EBIT']})",
         formula=f"=MIN(0,{{c}}{P['EBIT']}+{{p}}{r})")
    prow("Applicable tax rate", unit="15% then 25%", fmt=PCT,
         formula=(f"=IF({{c}}{P['Taxable profit']}<=0,0,"
                  f"IF(COUNTIF($C{P['Taxable profit']}:{{c}}{P['Taxable profit']},\">0\")"
                  f"<=Assumptions!{{c}}{A_ROW['tax_low_years']},"
                  f"Assumptions!{{c}}{A_ROW['tax_low']},Assumptions!{{c}}{A_ROW['tax_high']}))"))
    prow("Tax charge", formula=f"=-{{c}}{P['Taxable profit']}*{{c}}{r-1}")
    prow("NET PROFIT", formula=f"={{c}}{P['EBIT']}+{{c}}{r-1}", bold=True, top=True, band=True)

    # ══ CASH FLOW ═══════════════════════════════════════════════════════════
    cf = sheet(wb, "CashFlow", "Cash flow — indirect method")
    r, F = 4, {}

    def frow(label, formula=None, fmt=MONEY, **kw):
        nonlocal r
        F[label] = r
        r = row(cf, r, label, kw.pop("unit", ""), formula=formula, fmt=fmt, **kw)

    frow("Net profit", formula=f"=P&L!{{c}}{P['NET PROFIT']}", font=LINK)
    frow("Add back amortisation", formula=f"=WorkingCap!{{c}}{W['Amortisation']}", font=LINK)
    frow("Change in working capital", formula=f"=-WorkingCap!{{c}}{W['Change in NWC']}", font=LINK)
    frow("OPERATING CASH FLOW", bold=True, top=True,
         formula=f"=SUM({{c}}{r-3}:{{c}}{r-1})")
    frow("Capital expenditure", formula=f"=-WorkingCap!{{c}}{W['Capitalised development']}", font=LINK)
    frow("FREE CASH FLOW", bold=True, top=True, band=True,
         formula=f"={{c}}{F['OPERATING CASH FLOW']}+{{c}}{r-1}")
    frow("Cumulative free cash flow",
         first=f"={{c}}{F['FREE CASH FLOW']}",
         formula=f"={{p}}{r}+{{c}}{F['FREE CASH FLOW']}")
    r += 1
    frow("Equity raised", unit="see Funding",
         formula=f"=Funding!{{c}}{FUNDING_ROWS['Equity raised']}", font=LINK)
    # Year one opens at nothing rather than at `IF(k=1, 0, prev closing)`: in the
    # first column `{p}` falls back to the same column, so the untaken branch
    # named the closing-cash cell that reads this one. Excel does not care that
    # the branch never runs — it saw Opening -> Closing -> Opening and stopped
    # calculating. Found by the dependency-graph check, which the direct
    # self-reference check could not see because the loop goes through a
    # neighbour rather than back to the same cell.
    frow("Opening cash", first="=0", formula=f"={{p}}{r+1}")
    frow("CLOSING CASH", bold=True, top=True, band=True,
         formula=f"={{c}}{r-1}+{{c}}{F['FREE CASH FLOW']}+{{c}}{F['Equity raised']}")

    # ══ BALANCE SHEET ═══════════════════════════════════════════════════════
    bs = sheet(wb, "BalanceSheet", "Balance sheet")
    r, B = 4, {}

    def brow(label, formula=None, fmt=MONEY, **kw):
        nonlocal r
        B[label] = r
        r = row(bs, r, label, kw.pop("unit", ""), formula=formula, fmt=fmt, **kw)

    r = section(bs, r, "ASSETS")
    brow("Cash", formula=f"=CashFlow!{{c}}{F['CLOSING CASH']}", font=LINK)
    brow("Sponsor receivables", formula=f"=WorkingCap!{{c}}{W['Sponsor receivables']}", font=LINK)
    brow("Net intangible assets", formula=f"=WorkingCap!{{c}}{W['Net intangible assets']}", font=LINK)
    brow("TOTAL ASSETS", bold=True, top=True, band=True, formula=f"=SUM({{c}}{r-3}:{{c}}{r-1})")
    r += 1
    r = section(bs, r, "LIABILITIES")
    brow("Athlete payout float", formula=f"=WorkingCap!{{c}}{W['Athlete payout float']}", font=LINK)
    brow("Trade payables", formula=f"=WorkingCap!{{c}}{W['Trade payables']}", font=LINK)
    brow("TOTAL LIABILITIES", bold=True, top=True, formula=f"={{c}}{r-2}+{{c}}{r-1}")
    r += 1
    r = section(bs, r, "EQUITY")
    _eq = FUNDING_ROWS["Equity raised"]
    brow("Paid-in capital",
         first=f"=Funding!{{c}}{_eq}",
         formula=f"={{p}}{r}+Funding!{{c}}{_eq}")
    brow("Retained earnings",
         first=f"=P&L!{{c}}{P['NET PROFIT']}",
         formula=f"={{p}}{r}+P&L!{{c}}{P['NET PROFIT']}")
    brow("TOTAL EQUITY", bold=True, top=True, formula=f"={{c}}{r-2}+{{c}}{r-1}")
    brow("Liabilities + equity", bold=True,
         formula=f"={{c}}{B['TOTAL LIABILITIES']}+{{c}}{B['TOTAL EQUITY']}")
    brow("BALANCE CHECK", unit="must be zero", bold=True, top=True,
         formula=f"=ROUND({{c}}{B['TOTAL ASSETS']}-{{c}}{r-1},2)",
         font=Font(bold=True, color="B3272D", name="Calibri", size=10))

    # ══ FUNDING ═════════════════════════════════════════════════════════════
    fu = sheet(wb, "Funding", "Funding rounds and dilution")
    raises = [400_000, 0, 2_000_000, 0, 8_000_000, 0, 0, 0, 0, 0]
    pre = [2_500_000, 0, 10_000_000, 0, 40_000_000, 0, 0, 0, 0, 0]
    U = FUNDING_ROWS
    r = 4

    def urow(label, unit, **kw):
        nonlocal r
        assert r == U[label], f"Funding row moved: {label} is at {r}, declared {U[label]}"
        r = row(fu, r, label, unit, **kw)

    urow("Equity raised", "EUR", values=raises, fmt=MONEY)
    urow("Pre-money valuation", "EUR", values=pre, fmt=MONEY)
    urow("Post-money valuation", "EUR", fmt=MONEY,
         formula=f"=IF({{c}}{U['Equity raised']}=0,0,"
                 f"{{c}}{U['Pre-money valuation']}+{{c}}{U['Equity raised']})")
    urow("New investor stake", "%", fmt=PCT,
         formula=f"=IF({{c}}{U['Post-money valuation']}=0,0,"
                 f"{{c}}{U['Equity raised']}/{{c}}{U['Post-money valuation']})")
    urow("Cumulative dilution", "%", fmt=PCT,
         first=f"={{c}}{U['New investor stake']}",
         formula=f"=1-(1-{{p}}{U['Cumulative dilution']})"
                 f"*(1-{{c}}{U['New investor stake']})")
    urow("Founders + team retained", "%", fmt=PCT, bold=True,
         formula=f"=1-{{c}}{U['Cumulative dilution']}")

    # ══ VALUATION ═══════════════════════════════════════════════════════════
    va = sheet(wb, "Valuation", "Valuation — DCF, NPV, IRR and exit multiples")
    r = 4
    r = row(va, r, "Free cash flow", "EUR", formula=f"=CashFlow!{{c}}{F['FREE CASH FLOW']}",
            fmt=MONEY, font=LINK)
    r = row(va, r, "Discount factor", "1/(1+WACC)^t",
            formula=f"=1/(1+Assumptions!$C${A_ROW['wacc']})^{{k}}", fmt='0.000')
    r = row(va, r, "Discounted FCF", "EUR", formula="={c}4*{c}5", fmt=MONEY, bold=True)
    r += 1
    lastc = COLS[-1]
    # An FCF row with the terminal value added to the final year, so the
    # terminal-value IRR can be taken over a cash-flow series that actually
    # contains one. The previous formula ran to the second-to-last column and
    # included no terminal value at all, despite its label.
    r = row(va, r, "FCF incl. terminal value", "for the IRR below", fmt=MONEY,
            formula=f"={{c}}4+IF({{k}}={N},$C$__TV__,0)")
    tv_series_row = r - 1

    # Rows are recorded as they are written. The previous version addressed them
    # with arithmetic on an `r` captured before the loop ran, so the exit-value
    # formula pointed at its own output and Excel could not calculate it.
    V = {}
    blocks = [
        ("DCF", None),
        ("PV of explicit forecast", f"=SUM(C6:{lastc}6)"),
        ("Terminal value at Y10",
         f"={lastc}4*(1+Assumptions!$C${A_ROW['tg']})"
         f"/(Assumptions!$C${A_ROW['wacc']}-Assumptions!$C${A_ROW['tg']})"),
        ("PV of terminal value", "=C{Terminal value at Y10}*" + f"{lastc}5"),
        ("ENTERPRISE VALUE (DCF)", "=C{PV of explicit forecast}+C{PV of terminal value}"),
        ("", None),
        ("RETURN METRICS", None),
        ("NPV of FCF at WACC", f"=NPV(Assumptions!$C${A_ROW['wacc']},C4:{lastc}4)"),
        ("IRR of the plan", f"=IRR(C4:{lastc}4)"),
        ("IRR incl. terminal value", f"=IRR(C{tv_series_row}:{lastc}{tv_series_row},0.3)"),
        ("", None),
        ("EXIT MULTIPLE", None),
        ("Y10 net revenue", f"=Revenue!{lastc}{R['NET REVENUE']}"),
        ("Exit value at Y10", "=C{Y10 net revenue}*Assumptions!$C$" + str(A_ROW['exit_mult'])),
        ("Discounted to today", "=C{Exit value at Y10}*" + f"{lastc}5"),
    ]
    for label, formula in blocks:
        if formula is None:
            if label:
                r = section(va, r, label)
            else:
                r += 1
            continue
        V[label] = r
        va.cell(r, 1, label).font = BOLD if label.isupper() else Font(name="Calibri", size=10)
        cell = va.cell(r, 3, formula.format(**V))
        cell.number_format = PCT if "IRR" in label else MONEY
        cell.font = Font(bold=True, name="Calibri", size=10)
        r += 1
    # the terminal-value cell the helper row points at is only known now
    for k in range(N):
        c = va.cell(tv_series_row, FIRST + k)
        if isinstance(c.value, str):
            c.value = c.value.replace("$C$__TV__", f"$C${V['Terminal value at Y10']}")

    r += 1
    r = section(va, r, "SENSITIVITY — enterprise value by WACC and terminal growth")
    va.cell(r, 1, "Terminal growth \\ WACC").font = BOLD
    waccs = [0.18, 0.20, 0.22, 0.25, 0.28, 0.30]
    for j, w in enumerate(waccs):
        c = va.cell(r, 3 + j, w)
        c.number_format = PCT
        c.font = HEAD
        c.fill = HEAD_FILL
    r += 1
    for g in (0.01, 0.02, 0.03, 0.04, 0.05):
        va.cell(r, 1, g).number_format = PCT
        va.cell(r, 1).font = BOLD
        for j, w in enumerate(waccs):
            # PV of a 10-year FCF strip plus terminal value, at this (w, g) pair
            terms = "+".join(f"CashFlow!{c}{F['FREE CASH FLOW']}/(1+{w})^{k+1}"
                             for k, c in enumerate(COLS))
            tv = (f"CashFlow!{lastc}{F['FREE CASH FLOW']}*(1+{g})/({w}-{g})/(1+{w})^{N}")
            cell = va.cell(r, 3 + j, f"={terms}+{tv}")
            cell.number_format = MONEY
        r += 1



    # == COMPARABLES: published facts, nothing derived =======================
    cp = wb.create_sheet("Comparables")
    cp["A1"] = "Comparables - published facts about the platforms we are modelled on"
    cp["A1"].font = TITLE
    cp["A2"] = ("Nothing on this sheet is our estimate. Every figure is published and cited, so it can be "
                "refreshed independently of anything we concluded from it. MarketModel derives our "
                "assumptions from these numbers; Assumptions then uses what MarketModel produces.")
    cp["A2"].font = Font(italic=True, size=9, color="6B7480")
    for col, w in {"A": 38, "B": 16, "C": 16, "D": 15, "E": 13, "F": 46}.items():
        cp.column_dimensions[col].width = w

    r = 4
    for j, h in enumerate(["Metric", "Value", "Unit", "Platform", "Period", "Source"]):
        c = cp.cell(r, 1 + j, h); c.font, c.fill = HEAD, HEAD_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1
    CP = {}
    for metric, value, unit, platform, period, source in CD.PLATFORM_FACTS:
        CP[(platform, metric)] = r
        cp.cell(r, 1, metric).font = Font(size=10, name="Calibri")
        v = cp.cell(r, 2, value); v.fill, v.font = FILL_HARD, FONT_HARD
        v.number_format = '#,##0.000' if isinstance(value, float) and value < 100 else '#,##0'
        for j, t in ((3, unit), (4, platform), (5, period), (6, source)):
            cell = cp.cell(r, j, t)
            cell.font = Font(size=9, name="Calibri", color="4A525E")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for j in range(6):
            cp.cell(r, 1 + j).border = EDGE
        r += 1

    r += 1
    cp.cell(r, 1, "TAKE RATES - what each platform costs a creator").font = Font(bold=True, size=10, color="8A5200")
    r += 1
    for j, h in enumerate(["Platform", "Headline take", "Per txn (USD)", "Monthly fee (USD)", "", "Source"]):
        c = cp.cell(r, 1 + j, h); c.font, c.fill = HEAD, HEAD_FILL
    r += 1
    TK = {}
    for platform, take, per_txn, monthly, source in CD.TAKE_RATES:
        TK[platform] = r
        cp.cell(r, 1, platform).font = BOLD
        for j, (val, fmt) in enumerate([(take, PCT), (per_txn, MONEY2), (monthly, MONEY2)], start=2):
            c = cp.cell(r, j, val); c.fill, c.font, c.number_format = FILL_HARD, FONT_HARD, fmt
        cp.cell(r, 6, source).font = Font(size=9, color="4A525E", name="Calibri")
        r += 1

    r += 1
    cp.cell(r, 1, "INTERMEDIARIES the athlete already pays").font = Font(bold=True, size=10, color="8A5200")
    r += 1
    for name, lo, hi, source in CD.INTERMEDIARY_RATES:
        cp.cell(r, 1, name).font = Font(size=10, name="Calibri")
        for j, val in ((2, lo), (3, hi)):
            c = cp.cell(r, j, val); c.fill, c.font, c.number_format = FILL_HARD, FONT_HARD, PCT
        cp.cell(r, 6, source).font = Font(size=9, color="4A525E", name="Calibri")
        r += 1

    r += 2
    cp.cell(r, 1, "SOURCES").font = BOLD
    r += 1
    for label, url in CD.SOURCE_URLS.items():
        cp.cell(r, 1, label).font = Font(size=9, name="Calibri")
        c = cp.cell(r, 2, url)
        c.font = Font(size=9, color="1F4E9C", underline="single", name="Calibri")
        c.hyperlink = url
        cp.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    # == MARKET MODEL: derive our assumptions from the comparables ===========
    mm = wb.create_sheet("MarketModel")
    mm["A1"] = "Market Model - how the comparables become our assumptions"
    mm["A1"].font = TITLE
    mm["A2"] = ("Each block starts from a published figure on Comparables and ends in a number the "
                "Assumptions sheet uses. Adjustment factors are blue because they are our judgement; "
                "everything else is a formula you can trace.")
    mm["A2"].font = Font(italic=True, size=9, color="6B7480")
    for col, w in {"A": 46, "B": 16, "C": 14, "D": 62}.items():
        mm.column_dimensions[col].width = w
    legend(mm, 3)

    MM = {}
    r = 5

    def mrow(label, value=None, formula=None, note="", fmt=MONEY2, kind="calc", key=None):
        """kind: input (blue) | link (green, from Comparables) | calc | out (grey)"""
        nonlocal r
        MM[key or label] = r
        mm.cell(r, 1, label).font = BOLD if kind == "out" else Font(size=10, name="Calibri")
        c = mm.cell(r, 2)
        if formula is not None:
            c.value = formula
            c.fill = FILL_LINK if "Comparables!" in formula else (FILL_TOTAL if kind == "out" else FILL_CALC)
            c.font = FONT_LINK if "Comparables!" in formula else (BOLD if kind == "out" else FONT_CALC)
        else:
            c.value = value
            c.fill, c.font = FILL_INPUT, FONT_INPUT
        c.number_format = fmt
        c.border = EDGE
        n = mm.cell(r, 4, note)
        n.font = Font(size=9, color="4A525E", name="Calibri")
        n.alignment = Alignment(wrap_text=True, vertical="top")
        mm.row_dimensions[r].height = 30 if note else 15
        r += 1

    def sect(text):
        nonlocal r
        c = mm.cell(r, 1, text)
        c.font = Font(bold=True, size=10, color="8A5200")
        for j in range(4):
            mm.cell(r, 1 + j).fill = PatternFill("solid", fgColor="FFF4DE")
        r += 1

    # --- 1. paying fans per monetising creator ---
    sect("1. PAYING FANS PER MONETISING ATHLETE  -> Assumptions")
    mrow("Patreon active paying members", formula=f"=Comparables!B{CP[('Patreon','Active paying members')]}",
         fmt=NUM, note="Published")
    mrow("Patreon creators with >=1 paying member", formula=f"=Comparables!B{CP[('Patreon','Creators with >=1 paying member')]}",
         fmt=NUM, note="Published")
    mrow("Members per paying creator", formula=f"=B{MM['Patreon active paying members']}/B{MM['Patreon creators with >=1 paying member']}",
         fmt='0.0', note="The single most useful benchmark we found: it is measured, not modelled.",
         key="members_per_creator")
    mrow("OnlyFans fan accounts per creator",
         formula=f"=Comparables!B{CP[('OnlyFans','Fan accounts')]}/Comparables!B{CP[('OnlyFans','Creator accounts')]}",
         fmt='0.0', note="Fan ACCOUNTS, not paying fans - most follow free. Shown as an upper bound only.")
    mrow("Niche adjustment", MKT.NICHE_FPA_ADJUSTMENT, fmt='0.00', kind="input",
         note="OUR JUDGEMENT: +6% on the Patreon benchmark. Sport audiences are smaller but convert "
              "better, because the fan usually does the sport themselves.")
    mrow("Popular adjustment", MKT.POPULAR_FPA_ADJUSTMENT, fmt='0.00', kind="input",
         note="OUR JUDGEMENT: +37%. Much larger followings, much weaker conversion - more paying fans "
              "in absolute terms even though a smaller share converts.")
    mrow("→ Niche fans per athlete, mature", formula=f"=B{MM['members_per_creator']}*B{MM['Niche adjustment']}",
         fmt='0.0', kind="out", key="out_fpa_niche")
    mrow("→ Popular fans per athlete, mature", formula=f"=B{MM['members_per_creator']}*B{MM['Popular adjustment']}",
         fmt='0.0', kind="out", key="out_fpa_pop")

    # --- 2. fan ARPU from tier mix ---
    sect("2. FAN ARPU  -> Assumptions")
    mrow("Tier 1 price (Supporter)", MKT.TIER_PRICES[0], kind="input")
    mrow("Tier 2 price (Insider)", MKT.TIER_PRICES[1], kind="input")
    mrow("Tier 3 price (Inner circle)", MKT.TIER_PRICES[2], kind="input")
    mrow("Mix - Tier 1", MKT.TIER_MIX[0], fmt=PCT, kind="input", note="Assumed distribution; replace with the real tier mix after P1.")
    mrow("Mix - Tier 2", MKT.TIER_MIX[1], fmt=PCT, kind="input")
    mrow("Mix - Tier 3", MKT.TIER_MIX[2], fmt=PCT, kind="input")
    mrow("Mix check (must be 100%)",
         formula=f"=B{MM['Mix - Tier 1']}+B{MM['Mix - Tier 2']}+B{MM['Mix - Tier 3']}", fmt=PCT)
    mrow("→ Weighted ARPU per month",
         formula=(f"=B{MM['Tier 1 price (Supporter)']}*B{MM['Mix - Tier 1']}"
                  f"+B{MM['Tier 2 price (Insider)']}*B{MM['Mix - Tier 2']}"
                  f"+B{MM['Tier 3 price (Inner circle)']}*B{MM['Mix - Tier 3']}"),
         kind="out", key="out_arpu",
         note="Cross-check: Patreon reports $6.10 average support and a $8-12 typical band. "
              "Landing inside that band from an independent tier build is a good sign.")
    mrow("Patreon average support (cross-check)",
         formula=f"=Comparables!B{CP[('Patreon','Average monthly support per member')]}",
         note="Published average - below our figure because Patreon's long tail includes many $1-3 tiers. "
              "Our EUR 9.49 sits inside their $8-12 typical band.")
    mrow("Niche ARPU factor", MKT.NICHE_ARPU_FACTOR, fmt='0.00', kind="input",
         note="Niche fans buy knowledge and sit at the tier mix above.")
    mrow("Popular ARPU factor", MKT.POPULAR_ARPU_FACTOR, fmt='0.00', kind="input",
         note="OUR JUDGEMENT: -13%. Popular-sport fans skew to the cheapest tier.")
    mrow("→ Niche ARPU, mature",
         formula=f"=B{MM['out_arpu']}*B{MM['Niche ARPU factor']}", kind="out", key="out_arpu_niche")
    mrow("→ Popular ARPU, mature",
         formula=f"=B{MM['out_arpu']}*B{MM['Popular ARPU factor']}", kind="out", key="out_arpu_pop")

    # --- 3. fan churn ---
    sect("3. FAN CHURN  -> Assumptions")
    mrow("Patreon monthly churn (low)", formula=f"=Comparables!B{CP[('Patreon','Monthly churn (low)')]}", fmt=PCT)
    mrow("Patreon monthly churn (high)", formula=f"=Comparables!B{CP[('Patreon','Monthly churn (high)')]}", fmt=PCT)
    mrow("Benchmark midpoint",
         formula=f"=(B{MM['Patreon monthly churn (low)']}+B{MM['Patreon monthly churn (high)']})/2", fmt=PCT,
         key="churn_mid")
    mrow("Annual-plan churn multiplier", formula=f"=Comparables!B{CP[('Patreon','Annual-plan churn multiplier')]}",
         fmt='0.000', note="Patreon: annual patrons churn at one third the monthly rate.", key="ann_mult")
    mrow("Share of subscribers on annual plans", MKT.ANNUAL_PLAN_SHARE, fmt=PCT, kind="input",
         note="Our target for the season pass. This is the one lever that improves churn without "
              "changing the product.")
    mrow("Blended benchmark churn",
         formula=(f"=B{MM['churn_mid']}*(1-B{MM['Share of subscribers on annual plans']})"
                  f"+B{MM['churn_mid']}*B{MM['ann_mult']}*B{MM['Share of subscribers on annual plans']}"),
         fmt=PCT, key="churn_blend",
         note="What a Patreon-like platform would see with our annual mix.")
    mrow("Niche engagement factor", MKT.NICHE_ENGAGEMENT, fmt='0.00', kind="input",
         note="OUR CLAIM, UNTESTED, AND THE MOST OPTIMISTIC JUDGEMENT IN THE MODEL: this says niche "
              "fans churn 45% SLOWER than the Patreon benchmark, because training content is habitual "
              "and competitive seasons create renewal moments. Nothing yet proves it. If it is wrong "
              "and we are merely at benchmark, Y10 revenue falls by roughly EUR 7M.")
    mrow("Popular engagement factor", MKT.POPULAR_ENGAGEMENT, fmt='0.00', kind="input",
         note="OUR JUDGEMENT: 15% better than benchmark. Impulse follows after a result, with many "
              "free substitutes - but still a sport fan rather than a general creator audience.")
    mrow("→ Niche churn, mature", formula=f"=B{MM['churn_blend']}*B{MM['Niche engagement factor']}",
         fmt=PCT, kind="out", key="out_churn_niche")
    mrow("→ Popular churn, mature", formula=f"=B{MM['churn_blend']}*B{MM['Popular engagement factor']}",
         fmt=PCT, kind="out", key="out_churn_pop")

    # --- 4. creator earnings sanity check ---
    sect("4. WHAT AN ATHLETE ACTUALLY EARNS  (sanity check, not an input)")
    mrow("OnlyFans: paid to creators / creator accounts",
         formula=(f"=Comparables!B{CP[('OnlyFans','Paid to creators')]}"
                  f"/Comparables!B{CP[('OnlyFans','Creator accounts')]}/12"),
         note="Includes dormant accounts, so this is a floor.")
    mrow("Patreon: paid to creators / paying creators",
         formula=(f"=Comparables!B{CP[('Patreon','Paid to creators annually')]}"
                  f"/Comparables!B{CP[('Patreon','Creators with >=1 paying member')]}/12"),
         note="Denominator is creators WITH members, so this is a ceiling.")
    mrow("Stride: modelled niche athlete, mature",
         formula=f"=B{MM['out_fpa_niche']}*B{MM['out_arpu']}*(1-Assumptions!$C${A_ROW['take_fan']})",
         kind="out",
         note="Ours sits between the two, which is where a plausible figure should sit.")
    mrow("Revenue share taken by top 0.1% (OnlyFans)",
         formula=f"=Comparables!B{CP[('OnlyFans','Revenue concentration')]}", fmt=PCT,
         note="THE CAVEAT ON EVERY AVERAGE ABOVE: earnings follow a power law. Totals are unaffected, "
              "but the median athlete earns far less than the mean. Do not pitch the mean to an athlete.")

    # --- 5. competitive take rate ---
    sect("5. TAKE RATE - what an athlete keeps, by platform")
    mrow("USD to EUR", 0.92, fmt='0.00', kind="input",
         note="Comparables are published in USD; this model is in EUR. Applied to the per-transaction "
              "and monthly creator fees below.", key="fx")
    mrow("Athlete monthly fan revenue (test case)", 300.0, kind="input",
         note="Change this to see who is cheapest at any revenue level.", key="testrev")
    for plat in ("OnlyFans", "Patreon", "Passes", "Stride (proposed)"):
        row_ix = TK[plat]
        mrow(f"  {plat} - athlete keeps",
             formula=(f"=B{MM['testrev']}*(1-Comparables!B{row_ix})"
                      f"-B{MM['testrev']}/Assumptions!$C${A_ROW['fan_txn']}*Comparables!C{row_ix}*B{MM['fx']}"
                      f"-Comparables!D{row_ix}*B{MM['fx']}"),
             note="Headline take, plus per-transaction and monthly creator fees where they exist.")
    mrow("Crossover vs Passes (EUR/month)",
         formula=(f"=Comparables!D{TK['Passes']}*B{MM['fx']}/((Comparables!B{TK['Stride (proposed)']}"
                  f"-Comparables!B{TK['Passes']})-Comparables!C{TK['Passes']}*B{MM['fx']}"
                  f"/Assumptions!$C${A_ROW['fan_txn']})"),
         fmt=MONEY, kind="out",
         note="Below this, our flat 15% pays the athlete more than Passes' 10% plus fees. Our modelled "
              "athlete earns far less than this, so the long tail is on our side of the line.")

    # --- 6. TAM ---
    sect("6. MARKET SIZE  -> sanity-checks the athlete trajectory")
    total_pop = sum(CD.POPULATION_M.values())
    es_pop = CD.POPULATION_M["Spain"]
    mrow("Population, 34 countries in scope", total_pop, fmt='#,##0.0', kind="input", note="Millions. Eurostat / national statistics offices.")
    mrow("Population, Spain", es_pop, fmt='#,##0.0', kind="input", note="Millions.")
    mrow("Share who do any sport (Spain)", 0.56, fmt=PCT, kind="input",
         note="Eurobarometer 525 - Spain sits near the EU average of 55% who exercise at all.")
    mrow("Share with a public athletic identity", 0.02, fmt=PCT, kind="input",
         note="OUR ESTIMATE: competes, posts publicly, has a following worth sponsoring.")
    mrow("Share with commercial intent", 0.25, fmt=PCT, kind="input",
         note="OUR ESTIMATE: would monetise if the tooling existed.")
    mrow("→ Addressable athletes, Spain",
         formula=(f"=B{MM['Population, Spain']}*1000000*B{MM['Share who do any sport (Spain)']}"
                  f"*B{MM['Share with a public athletic identity']}*B{MM['Share with commercial intent']}"),
         fmt=NUM, kind="out", key="sam_es")
    mrow("→ Addressable athletes, all 34 countries",
         formula=(f"=B{MM['Population, 34 countries in scope']}*1000000*0.5"
                  f"*B{MM['Share with a public athletic identity']}*B{MM['Share with commercial intent']}"),
         fmt=NUM, kind="out", key="tam_all",
         note="Using a 50% blended participation rate across the 34 countries.")
    mrow("Y10 athletes in the plan", formula=f"=Assumptions!${COLS[-1]}${A_ROW['athletes']}", fmt=NUM)
    mrow("→ Y10 penetration of Spain SAM",
         formula=f"=Assumptions!${COLS[-1]}${A_ROW['athletes']}/B{MM['sam_es']}", fmt=PCT, kind="out")
    mrow("→ Y10 penetration of full TAM",
         formula=f"=Assumptions!${COLS[-1]}${A_ROW['athletes']}/B{MM['tam_all']}", fmt=PCT, kind="out",
         note="If this reads as implausibly high, the athlete trajectory is too aggressive - that is "
              "exactly what this block exists to reveal.")

    # == RESEARCH & BENCHMARKING =============================================
    rs = wb.create_sheet("Research")
    rs["A1"] = "Research & Benchmarking - where every assumption came from"
    rs["A1"].font = TITLE
    for col, w in {"A": 38, "B": 12, "C": 14, "D": 54, "E": 44, "F": 11, "G": 44}.items():
        rs.column_dimensions[col].width = w
    rs["A2"] = ("SOURCED = published figure | BENCHMARKED = set against named comparables | "
                "DERIVED = computed from other assumptions | ESTIMATE = judgement, no benchmark yet")
    rs["A2"].font = Font(italic=True, size=9, color="6B7480")

    for j, h in enumerate(["Assumption", "Model value", "Method", "Benchmark / comparable",
                           "Source", "Confidence", "What would replace the estimate"]):
        c = rs.cell(4, 1 + j, h)
        c.font, c.fill = HEAD, HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rs.freeze_panes = "A5"

    conf_fill = {"High": FILL_LINK, "Medium": FILL_HARD, "Low": FILL_CHECK}
    method_col = {"SOURCED": "1E7A3C", "BENCHMARKED": "1F4E9C",
                  "DERIVED": "1A1A1A", "ESTIMATE": "B3272D"}

    r = 5
    for entry in RD.ROWS:
        if len(entry) == 1:
            c = rs.cell(r, 1, entry[0])
            c.font = Font(bold=True, size=10, color="8A5200")
            for j in range(7):
                rs.cell(r, 1 + j).fill = PatternFill("solid", fgColor="FFF4DE")
            r += 1
            continue
        label, key, method, bench, source, conf, improve = entry
        rs.cell(r, 1, label).font = BOLD
        if key and key in A_ROW:
            cell = rs.cell(r, 2, "=Assumptions!$C$%d" % A_ROW[key])
            cell.fill, cell.font, cell.number_format = FILL_LINK, FONT_LINK, "0.000"
        else:
            rs.cell(r, 2, "see model").font = Font(size=9, italic=True, color="6B7480")
        m = rs.cell(r, 3, method)
        m.font = Font(bold=True, size=9, color=method_col[method])
        for j, text in ((4, bench), (5, source), (7, improve)):
            c = rs.cell(r, j, text)
            c.font = Font(size=9, name="Calibri")
            c.alignment = Alignment(wrap_text=True, vertical="top")
        cf = rs.cell(r, 6, conf)
        cf.fill, cf.font = conf_fill[conf], Font(bold=True, size=9)
        cf.alignment = Alignment(horizontal="center")
        for j in range(7):
            rs.cell(r, 1 + j).border = EDGE
        rs.row_dimensions[r].height = 58
        r += 1

    r += 1
    c = rs.cell(r, 1, "THE THREE ASSUMPTIONS MOST LIKELY TO BE WRONG")
    c.font = Font(bold=True, size=11, color="B3272D")
    r += 1
    for text in RD.MOST_LIKELY_WRONG:
        cell = rs.cell(r, 1, text)
        cell.font = Font(size=9, italic=not text[:2].strip().rstrip(".").isdigit())
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        rs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        rs.row_dimensions[r].height = 30
        r += 1

    # ══ CHECK ═══════════════════════════════════════════════════════════════
    ck = sheet(wb, "Check", "Check — Excel's own answers against the Python model")
    r = 4
    ck.cell(2, 1, "Each pair is the Python model's figure and the workbook's own calculation. "
                  "VARIANCE must be zero — if a formula is wrong, it shows up here on open.").font =         Font(italic=True, size=9, color="6B7480")

    CHECKS = [
        ("Net revenue", "revenue", f"=Revenue!{{c}}{R['NET REVENUE']}", MONEY),
        ("Cost of sales", "cogs", f"=Costs!{{c}}{C['TOTAL COST OF SALES']}", MONEY),
        ("Gross profit", "gross", f"=P&L!{{c}}{P['GROSS PROFIT']}", MONEY),
        ("Operating costs", "opex", f"=Costs!{{c}}{C['TOTAL OPERATING COSTS']}", MONEY),
        ("EBITDA", "ebitda", f"=P&L!{{c}}{P['EBITDA']}", MONEY),
        ("Paying fans (year end)", "paying_fans", f"=Drivers!{{c}}{D['Paying fans (year end)']}", NUM),
        ("Average paying fans", "avg_fans", f"=Drivers!{{c}}{D['Average paying fans']}", NUM),
        ("Total GMV", "gmv", f"=Revenue!{{c}}{R['TOTAL GMV']}", MONEY),
        ("Deals", "deals", f"=Drivers!{{c}}{D['Total deals']}", NUM),
        ("Applications required", "applications", f"=Drivers!{{c}}{D['Applications required']}", NUM),
        ("Athlete verification", "verification", f"=Costs!{{c}}{C['Athlete verification']}", MONEY),
        # The three lines below the EBITDA line. Leaving them out is how the two
        # models drifted apart unnoticed: Python taxed a flat 15% forever and
        # called EBITDA-less-tax "free cash flow", while the workbook ran a loss
        # carry-forward, a 15%->25% step, working capital and capex. Both were
        # internally consistent and they disagreed, and nothing compared them.
        ("Tax charge", "tax", f"=-P&L!{{c}}{P['Tax charge']}", MONEY),
        ("Net profit", "net_profit", f"=P&L!{{c}}{P['NET PROFIT']}", MONEY),
        ("Free cash flow", "fcf", f"=CashFlow!{{c}}{F['FREE CASH FLOW']}", MONEY),
    ]
    for label, key, formula, fmt in CHECKS:
        r = section(ck, r, label.upper())
        py = r
        r = row(ck, r, "  Python model", "model.py", values=[x[key] for x in rows], fmt=fmt)
        xl = r
        r = row(ck, r, "  This workbook", "recalculated", formula=formula, fmt=fmt)
        r = row(ck, r, "  VARIANCE", "must be 0", fmt=fmt, check=True, bold=True,
                formula=f"=ROUND({{c}}{xl}-{{c}}{py},0)")
        r += 1

    ck.cell(r, 1, "Balance sheet check (from BalanceSheet, must be zero):").font = BOLD
    r = row(ck, r + 1, "  Assets less liabilities and equity", "must be 0", check=True, bold=True,
            formula=f"=BalanceSheet!{{c}}{B['BALANCE CHECK']}", fmt=MONEY)

    out = pathlib.Path(__file__).parent / "Stride_Financial_Model.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    path = build()
    size = path.stat().st_size / 1024
    print(f"wrote {path.name} ({size:.0f} KB)")
