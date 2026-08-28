"""Static checks on the generated workbook, for a machine with no Excel on it.

    python scripts/verify_workbook.py

The workbook is the deliverable a reader opens in Excel, and this repository
cannot open it. Each version of this script has been written after the previous
one passed a file that was broken:

  * v1 confirmed every reference pointed at a cell holding something. It did,
    and forty formulas still said `=P&L!C21`, which Excel cannot parse at all
    because the sheet token ends at the ampersand.
  * v2 added quoting and self-reference checks, and found six circular formulas
    the same afternoon — but only *direct* ones, `C8` naming `C8`. A cycle that
    goes through another cell, or another sheet, walked straight past it.

So the circularity check is now a real dependency graph over every cell, and a
self-reference is simply the shortest cycle in it.

    1. DANGLING   a reference to a cell that holds nothing — a silent zero
    2. QUOTING    a sheet name Excel cannot read bare, unquoted
    3. CYCLES     any closed loop of references, of any length, across sheets;
                  Excel stops calculating and shows 0
    4. SYNTAX     anything the formula tokenizer cannot read

Not covered, and worth knowing: `OFFSET` and other functions that compute a
reference at runtime cannot be resolved statically, so their dependencies are
invisible here. The one use of it in this workbook is amortisation, which reads
backwards along a row and cannot close a loop.
"""

from __future__ import annotations

import pathlib
import re
import sys

from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils import column_index_from_string, get_column_letter

WORKBOOK = (pathlib.Path(__file__).resolve().parents[1]
            / "business-plan" / "Stride_Financial_Model.xlsx")

BARE_SHEET = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
SHEET_PREFIX = r"(?:'(?P<q>[^']+)'|(?P<b>[A-Za-z_][A-Za-z0-9_.&]*))!"
# a range or a single cell, optionally sheet-qualified
REF = re.compile(
    rf"(?:{SHEET_PREFIX})?\$?(?P<c1>[A-Z]{{1,3}})\$?(?P<r1>\d{{1,5}})"
    rf"(?::\$?(?P<c2>[A-Z]{{1,3}})\$?(?P<r2>\d{{1,5}}))?(?![0-9(])")

Cell = tuple[str, int, int]      # sheet, row, column


def references(formula: str, here_sheet: str, sheets: set[str]):
    """Every cell a formula reads, ranges expanded."""
    for m in REF.finditer(formula):
        named = m.group("q") or m.group("b")
        if named and named not in sheets:
            continue                       # a function name, not a sheet
        sheet = named or here_sheet
        c1, r1 = column_index_from_string(m.group("c1")), int(m.group("r1"))
        if m.group("c2"):
            c2, r2 = column_index_from_string(m.group("c2")), int(m.group("r2"))
            for col in range(min(c1, c2), max(c1, c2) + 1):
                for rw in range(min(r1, r2), max(r1, r2) + 1):
                    yield (sheet, rw, col)
        else:
            yield (sheet, r1, c1)


def show(cell: Cell) -> str:
    sheet, rw, col = cell
    return f"{sheet}!{get_column_letter(col)}{rw}"


def find_cycle(graph: dict[Cell, set[Cell]]) -> list[Cell] | None:
    """First cycle found, as the path around it. Iterative so a deep chain of
    references cannot blow the stack on a large sheet."""
    WHITE, GREY, BLACK = 0, 1, 2
    colour: dict[Cell, int] = {}
    for start in graph:
        if colour.get(start, WHITE) != WHITE:
            continue
        stack = [(start, iter(graph.get(start, ())))]
        path = [start]
        colour[start] = GREY
        while stack:
            node, children = stack[-1]
            advanced = False
            for child in children:
                state = colour.get(child, WHITE)
                if state == GREY:
                    return path[path.index(child):] + [child]
                if state == WHITE:
                    colour[child] = GREY
                    path.append(child)
                    stack.append((child, iter(graph.get(child, ()))))
                    advanced = True
                    break
            if not advanced:
                colour[node] = BLACK
                stack.pop()
                path.pop()
    return None


def main() -> int:
    if not WORKBOOK.exists():
        print(f"no workbook at {WORKBOOK} — run business-plan/build_workbook.py first")
        return 1
    wb = load_workbook(WORKBOOK)
    sheets = set(wb.sheetnames)
    problems: list[str] = []
    graph: dict[Cell, set[Cell]] = {}
    formulas = refs = 0

    for ws in wb.worksheets:
        for line in ws.iter_rows():
            for cell in line:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                formulas += 1
                here = f"{ws.title}!{cell.coordinate}"

                try:
                    Tokenizer(value)
                except Exception as exc:
                    problems.append(f"SYNTAX   {here}: {exc}  [{value[:70]}]")
                    continue

                for m in REF.finditer(value):
                    bare = m.group("b")
                    if bare and bare not in sheets and not BARE_SHEET.match(bare):
                        problems.append(
                            f"QUOTING  {here}: sheet {bare!r} must be quoted  [{value[:70]}]")

                node: Cell = (ws.title, cell.row, cell.column)
                edges = graph.setdefault(node, set())
                for target in references(value, ws.title, sheets):
                    refs += 1
                    edges.add(target)
                    sheet, rw, col = target
                    if wb[sheet].cell(rw, col).value is None:
                        problems.append(f"DANGLING {here} -> empty {show(target)}")

    cycle = find_cycle(graph)
    if cycle:
        problems.append("CYCLE    " + " -> ".join(show(c) for c in cycle))

    print(f"{WORKBOOK.name}: {formulas:,} formulas, {refs:,} references, "
          f"{len(graph):,} nodes in the dependency graph")
    if problems:
        kinds: dict[str, int] = {}
        for pr in problems:
            kinds[pr.split()[0]] = kinds.get(pr.split()[0], 0) + 1
        print(f"\n{len(problems)} problems: "
              + ", ".join(f"{k} {v}" for k, v in sorted(kinds.items())))
        for pr in problems[:25]:
            print("  -", pr)
        if len(problems) > 25:
            print(f"  ... and {len(problems) - 25} more")
        return 1
    print("no dangling references, no unquoted sheet names, no cycles of any "
          "length, every formula parses.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
